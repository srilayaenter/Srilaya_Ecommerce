"""
SriLaYa Cost Sheet — Full Architecture
Sheet 1: Assumptions     — packaging/label/freight/labour/margins (user fills yellow cells)
Sheet 2: Laddu Recipes   — per-variety ingredient ratios → RM cost per kg
Sheet 3: Cost Model      — all 199 variants with formula-driven landed cost + suggested price
"""
import json, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Read selling prices from existing master Excel (col Q) ───────────────────
_OUT = r"D:\CompanyWebsite\SriLaYa_Cost_Sheet.xlsx"
SAVED_PRICES = {}
if os.path.exists(_OUT):
    try:
        _wb = openpyxl.load_workbook(_OUT, data_only=True)
        _ws = _wb["Cost Model"]
        for _r in range(4, 300):
            _sku = _ws.cell(_r, 3).value
            _q   = _ws.cell(_r, 17).value
            if _sku and _q is not None:
                SAVED_PRICES[_sku] = _q
        _wb.close()
        print(f"Loaded {len(SAVED_PRICES)} saved prices from master Excel")
    except Exception as e:
        print(f"Warning: could not read master Excel ({e}), using fallback prices")

# ── Evenmore 5kg tier rates (₹/kg) ───────────────────────────────────────────
EM = {
    "Foxtail Flakes": 106, "Little Flakes": 191, "Kodo Flakes": 116,
    "Barnyard Flakes": 197, "Ragi Flakes": 82, "Pearl Flakes": 79,
    "White Sorghum Flakes": 80, "Red Sorghum Flakes": 88, "Wheat Flakes": 94,
    "Barley Flakes": 92, "Mapillai Samba Flakes": 95, "Karupukavuni Flakes": 147,
    "Poongar Flakes": 100, "Kattuyanam Flakes": 100, "Karunguruvai Flakes": 98,
    "Thanga Samba Flakes": 93, "Green Gram Flakes": 151, "Horsegram Flakes": 93,
    "Red Matta Rice Flakes": 89, "Red Rice Flakes": 93,
    "Foxtail Rava": 105, "Little Rava": 216, "Barnyard Rava": 222,
    "Finger Rava": 79, "Pearl Rava": 70, "White Sorghum Rava": 72, "Mapillai Samba Rava": 111,
    "Foxtail Millet Grains": 68, "Little Millet Grains": 140, "Kodo Millet Grains": 80,
    "Barnyard Millet Grains": 143, "Finger Millet Grains": 54, "Pearl Millet Grains": 45,
    "White Sorghum Grains": 48, "Red Sorghum Grains": 53,
    "Parboiled Foxtail": 99.65, "Parboiled Little": 176.89, "Parboiled Kodo": 109.94,
    "Parboiled Barnyard": 178.95, "Parboiled Ragi": 80.07, "Parboiled Pearl": 70.29,
    "Parboiled White Sorghum": 70.80,
    "Foxtail Flour": 81, "Little Flour": 148, "Barnyard Flour": 149,
    "Finger Flour": 59, "Pearl Flour": 49, "White Sorghum Flour": 50,
    "Sprouted Ragi Flour": 81, "Whole Wheat Flour": 51, "Mapillai Samba Flour": 77,
    "Honey Muesli": 265, "Chocolate Muesli": 298, "Strawberry Muesli": 305,
    "Fruits Granola": 270, "Mixed Fruit and Nuts Muesli": 264,
    "Sugar Free Muesli": 246, "Sugar Free Stevia Muesli": 282,
    "Sugarcane Jaggery": 66, "Coconut Jaggery": 321, "Palm Jaggery": 328,
    "Unrefined Cane Sugar": 66,
    "Health Mix": 170, "Ragi Malt Carrot": 214, "Ragi Malt Beetroot": 207,
    "Mapillai Samba Rice": 69, "Karupu Kavuni Rice": 124, "Poongar Rice": 73,
}

# ── Per-SKU RM rate overrides (from SriLaYa_RM_Rates_Input.xlsx) ─────────────
SKU_RATES = {
  "WSFL-500G": 50,  "WSFL-1KG": 50,
  "RGVA-500G": 80,  "RGVA-1KG": 80,
  "WSVA-500G": 72,  "WSVA-1KG": 72,
  "BARI-500G": 66,  "BARI-1KG": 66,
  "BYR-500G":  143, "BYR-1KG":  143,
  "BTR-500G":  142, "BTR-1KG":  142,
  "FXR-500G":  68,  "FXR-1KG":  68,
  "KDR-500G":  80,  "KDR-1KG":  80,
  "LTR-500G":  140, "LTR-1KG":  140,
  "PLR-500G":  45,  "PLR-1KG":  45,
  "RSR-500G":  53,  "RSR-1KG":  53,
  "RGR-500G":  54,  "RGR-1KG":  54,
  "WSR-500G":  48,  "WSR-1KG":  48,
  "BPR-500G":  179, "BPR-1KG":  179,
  "FPR-500G":  100, "FPR-1KG":  100,
  "KPR-500G":  110, "KPR-1KG":  110,
  "LPR-500G":  177, "LPR-1KG":  177,
  "PPR-500G":  71,  "PPR-1KG":  71,
  "RPR-500G":  81,  "RPR-1KG":  81,
  "WPR-500G":  71,  "WPR-1KG":  71,
}

def em_rate(title):
    if title in EM: return EM[title]
    tl = title.lower()
    for k, v in EM.items():
        if k.lower() in tl or tl in k.lower(): return v
    return None

# ── Laddu recipes from ops app (per 1 kg of finished product) ────────────────
LADDU_RECIPES = {
    "Groundnut Laddu":           {"Groundnut": 500, "Naattu Sakkarai": 375, "Pottukadalai": 125},
    "Barnyard Millet Laddu":     {"Barnyard Millet": 300, "Groundnut": 300, "Naattu Sakkarai": 300, "Ghee": 100},
    "Foxtail Millet Laddu":      {"Foxtail Millet": 380, "Moong Dal": 190, "Naattu Sakkarai": 380, "Ghee": 50},
    "Till & Little Millet Laddu":{"Little Millet": 400, "Till (Sesame)": 200, "Naattu Sakkarai": 350, "Ghee": 50},
    "Till Sesame Laddu":         {"Till (Sesame)": 370, "Naattu Sakkarai": 490, "Pottukadalai": 120, "Ghee": 20},
    "Green Gram Laddu":          {"Green Gram": 460, "Naattu Sakkarai": 490, "Ghee": 50},
    "Horse Gram Laddu":          {"Horse Gram": 375, "Moong Dal": 188, "Pottukadalai": 188, "Naattu Sakkarai": 222, "Ghee": 25},
    "Moth Bean Laddu":           {"Moth Bean": 250, "Pottukadalai": 250, "Naattu Sakkarai": 500, "Ghee": 30},
    "Black Urad Dal Laddu":      {"Black Urad Dal": 281, "Groundnut": 188, "Pottukadalai": 188, "Naattu Sakkarai": 281, "Rice": 47, "Ghee": 47},
    "Pearl Millet Laddu":        {"Pearl Millet": 500, "Groundnut": 125, "Naattu Sakkarai": 250, "Ghee": 78},
}

ALL_INGREDIENTS = [
    "Barnyard Millet", "Foxtail Millet", "Little Millet", "Pearl Millet",
    "Groundnut", "Till (Sesame)", "Pottukadalai",
    "Naattu Sakkarai", "Ghee",
    "Green Gram", "Horse Gram", "Moth Bean",
    "Black Urad Dal", "Moong Dal", "Rice",
    "Other Ingredient",
]

# ── Styles ────────────────────────────────────────────────────────────────────
G_DARK  = "006A38"
G_MID   = "2E6F40"
G_LIGHT = "E8F5EE"
YELLOW  = "FFFF00"
AMBER   = "FFF3CD"
BLUE_T  = "EBF3FB"
WHITE   = "FFFFFF"
RED_HDR = "AA2222"

def cb():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_b():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def hf(c): return PatternFill("solid", fgColor=c)

def bf(sz=11, col="000000", bold=False, italic=False):
    return Font(name="Arial", size=sz, color=col, bold=bold, italic=italic)

def hdr(ws, row, col, val, bg=G_DARK, fg=WHITE, sz=11, span=1, height=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = bf(sz, fg, bold=True)
    c.fill = hf(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = cb()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    if height:
        ws.row_dimensions[row].height = height
    return c

def inp(ws, row, col, val=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = hf(YELLOW)
    c.font = bf(11, "0000FF", bold=True)
    c.border = thick_b()
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: c.number_format = fmt
    return c

def lbl(ws, row, col, val, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = bf(11, "000000", bold=bold)
    c.border = cb()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def nt(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = bf(9, "666666", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return c

def section(ws, row, title, ncols=4):
    c = ws.cell(row=row, column=1, value=title)
    c.font = bf(10, WHITE, bold=True)
    c.fill = hf(G_DARK)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 18

wb = openpyxl.Workbook()
wb.calculation.calcMode = 'auto'
wb.calculation.fullCalcOnLoad = True

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Assumptions"
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 6
ws1.column_dimensions["D"].width = 38

hdr(ws1, 1, 1, "SriLaYa — Cost Assumptions  (fill YELLOW cells; everything else auto-updates)",
    span=4, height=28, sz=13)
nt(ws1, 2, 1, "Fill these yellow cells once — the Cost Model sheet recalculates automatically.")
ws1.merge_cells("A2:D2")

# A. TRANSPORT
section(ws1, 4, "A. Transport / Inbound Freight")
rows_A = [
    ("Freight per kg — Evenmore shipments (₹/kg)",  5,  '₹#,##0.00', "Total shipping cost ÷ kg ordered"),
    ("Loading / handling / misc per kg (₹/kg)",     1,  '₹#,##0.00', "Unloading labour, transit losses"),
]
for i, (label, dflt, fmt, n) in enumerate(rows_A):
    r = 5 + i
    lbl(ws1, r, 1, label); inp(ws1, r, 2, dflt, fmt); nt(ws1, r, 4, n)

# B. POUCH PACKAGING  (rows 10–14)
section(ws1, 9, "B. Packaging — Pouches (cost per pouch from your supplier)")
hdr(ws1, 10, 1, "Size", G_MID, WHITE, sz=10)
hdr(ws1, 10, 2, "Cost/pouch (₹)", G_MID, WHITE, sz=10)
hdr(ws1, 10, 4, "Note", G_MID, WHITE, sz=10)
pouches = [
    ("200g stand-up pouch", 2.5, "Small pouch"),
    ("250g stand-up pouch", 2.5, "Same mould as 200g"),
    ("500g stand-up pouch", 3.5, "Medium pouch"),
    ("1kg  stand-up pouch", 5.0, "Large flakes pouch (160×120mm)"),
    ("2kg  stand-up pouch", 7.0, "Jumbo pouch"),
]
for i, (sz, dflt, n) in enumerate(pouches):
    r = 11 + i
    lbl(ws1, r, 1, sz); inp(ws1, r, 2, dflt, '₹#,##0.00'); nt(ws1, r, 4, n)
# Pouch rows: B11=200g, B12=250g, B13=500g, B14=1kg, B15=2kg

# C. BOX PACKAGING (Laddu)  (rows 18–20)
section(ws1, 17, "C. Packaging — Laddu Boxes (cost per box)")
hdr(ws1, 18, 1, "Size", G_MID, WHITE, sz=10)
hdr(ws1, 18, 2, "Cost/box (₹)", G_MID, WHITE, sz=10)
hdr(ws1, 18, 4, "Note", G_MID, WHITE, sz=10)
boxes = [
    ("250g thick box (Laddu)", 8,  "Rigid gift-style box"),
    ("500g thick box (Laddu)", 12, "Rigid gift-style box"),
    ("1kg  thick box (Laddu)", 18, "Rigid gift-style box"),
]
for i, (sz, dflt, n) in enumerate(boxes):
    r = 19 + i
    lbl(ws1, r, 1, sz); inp(ws1, r, 2, dflt, '₹#,##0.00'); nt(ws1, r, 4, n)
# Box rows: B19=250g, B20=500g, B21=1kg

# D. LABEL & FINISHING  (rows 23–28)
section(ws1, 22, "D. Label & Finishing (per unit)")
rows_D = [
    ("Label — pouches (₹/unit)",                    2.5, '₹#,##0.00', "Print + cut cost"),
    ("Label — laddu boxes (₹/unit)",                2.5, '₹#,##0.00', "Box may use a larger label"),
    ("Heat-seal finishing — pouches (₹/unit)",       0.5, '₹#,##0.00', "Electricity + sealer wear"),
    ("Box finishing: tape/insert/ribbon (₹/unit)",   1.0, '₹#,##0.00', "Tissue, sticker, ribbon if used"),
    ("Wastage allowance — pouches (%)",           0.02,   '0.00%',     "% RM lost in weighing + filling — enter as 0.02 for 2%"),
    ("Wastage allowance — laddus (%)",            0.03,   '0.00%',     "Moisture loss + breakage — enter as 0.03 for 3%"),
]
for i, (label, dflt, fmt, n) in enumerate(rows_D):
    r = 23 + i
    lbl(ws1, r, 1, label); inp(ws1, r, 2, dflt, fmt); nt(ws1, r, 4, n)
# D rows: B23=label pouch, B24=label box, B25=seal pouch, B26=box finishing, B27=wastage pouch%, B28=wastage laddu%

# E. LABOUR  (rows 30–32)
section(ws1, 29, "E. Labour")
rows_E = [
    ("Labour rate (₹/hour/person)",             175,  '₹#,##0.00', "₹1,400/day ÷ 8 hrs = ₹175/hr"),
    ("Laddu production rate (kg finished/hr)",    1,  '0.0',       "Your info: 1 kg per person-hour"),
    ("Pouch filling rate (units/hr/person)",     30,  '0',         "Weigh + fill + seal per hour"),
]
for i, (label, dflt, fmt, n) in enumerate(rows_E):
    r = 30 + i
    lbl(ws1, r, 1, label); inp(ws1, r, 2, dflt, fmt); nt(ws1, r, 4, n)
# E rows: B30=labour rate, B31=laddu rate, B32=pouch rate

# F. LADDU INGREDIENT RATES  (rows 35–45)
section(ws1, 34, "F. Laddu Ingredient Rates (₹/kg — your supplier prices)")
nt(ws1, 35, 1, "Naattu Sakkarai + millets pre-filled from Evenmore 5kg rate. Fill the rest from your purchases.")
ws1.merge_cells("A35:D35")
hdr(ws1, 36, 1, "Ingredient", G_MID, WHITE, sz=10)
hdr(ws1, 36, 2, "Rate (₹/kg)", G_MID, WHITE, sz=10)
hdr(ws1, 36, 4, "Source", G_MID, WHITE, sz=10)
ingredients_F = [
    ("Naattu Sakkarai (Sugarcane Jaggery)",  66,   "Evenmore 5kg bulk rate"),
    ("Groundnut (raw, shelled)",             150,  "Local supplier"),
    ("Pottukadalai (Roasted Chickpea/Gram)", 130,  "Local supplier"),
    ("Ghee (cow / buffalo)",                 540,  "Local supplier / bulk rate"),
    ("Till / Sesame seeds",                  600,  "Local supplier"),
    ("Green Gram (Moong)",                   140,  "Local supplier"),
    ("Horse Gram (Kollu)",                    90,  "Local supplier"),
    ("Moth Bean (Naripayiru)",               120,  "Local supplier"),
    ("Barnyard Millet grains",               143,  "Evenmore 5kg rate"),
    ("Foxtail Millet grains",                 68,  "Evenmore 5kg rate"),
    ("Little Millet grains",                 140,  "Evenmore 5kg rate"),
    ("Pearl Millet grains",                   45,  "Evenmore 5kg rate"),
    ("Black Urad Dal",                       140,  "Local supplier"),
    ("Moong Dal",                            130,  "Local supplier"),
    ("Rice (raw)",                            90,  "Local supplier"),
]
INGR_ROWS = {}
for i, (ing, dflt, n) in enumerate(ingredients_F):
    r = 37 + i
    lbl(ws1, r, 1, ing)
    inp(ws1, r, 2, float(dflt) if dflt is not None else None, '₹#,##0.00')
    nt(ws1, r, 4, n)
    INGR_ROWS[ing] = r
# F rows: B37=NS, B38=Groundnut, B39=Pottukadalai, B40=Ghee, B41=Till, B42=GreenGram,
# B43=HorseGram, B44=MothBean, B45=Barnyard, B46=Foxtail, B47=Little,
# B48=Pearl, B49=BlackUradDal, B50=MoongDal, B51=Rice  (last ingredient row = 51)

# G. TARGET MARGINS  (rows 53–56 — pushed down to avoid overlap with F ingredients)
section(ws1, 53, "G. Target Retail Margins (% above landed cost)")
rows_G = [
    ("Conservative margin %", 0.35, "Minimum — covers costs + modest profit — enter as 0.35 for 35%"),
    ("Target margin %",       0.45, "Recommended for most products — enter as 0.45 for 45%"),
    ("Premium margin %",      0.55, "Specialty / heritage / Laddu — enter as 0.55 for 55%"),
]
for i, (label, dflt, n) in enumerate(rows_G):
    r = 54 + i
    lbl(ws1, r, 1, label); inp(ws1, r, 2, dflt, '0%'); nt(ws1, r, 4, n)
# G rows: B54=conservative, B55=target, B56=premium

# H. GST REFERENCE
section(ws1, 58, "H. GST Reference (read-only — verify with CA)")
gst = [
    ("Millets, Rice, Rava, Flour, Flakes (branded, packaged)", "5%",  "Likely applicable; confirm"),
    ("Muesli, Granola, Health Mixes",                          "18%", "Processed breakfast cereal"),
    ("Sweeteners — Jaggery (all types)",                       "0%",  "Unprocessed jaggery = NIL"),
    ("Laddu (homemade sweet)",                                  "5%",  "Confirm with CA"),
]
for i, (cat, rate, n) in enumerate(gst):
    r = 59 + i
    lbl(ws1, r, 1, cat)
    c = ws1.cell(row=r, column=2, value=rate)
    c.font = bf(11, bold=True); c.border = cb()
    nt(ws1, r, 4, n)

# ── ingredient → Assumptions row lookup ──────────────────────────────────────
ING_RATE = {
    "Naattu Sakkarai": f"Assumptions!$B${INGR_ROWS['Naattu Sakkarai (Sugarcane Jaggery)']}",
    "Groundnut":       f"Assumptions!$B${INGR_ROWS['Groundnut (raw, shelled)']}",
    "Pottukadalai":    f"Assumptions!$B${INGR_ROWS['Pottukadalai (Roasted Chickpea/Gram)']}",
    "Ghee":            f"Assumptions!$B${INGR_ROWS['Ghee (cow / buffalo)']}",
    "Till (Sesame)":   f"Assumptions!$B${INGR_ROWS['Till / Sesame seeds']}",
    "Green Gram":      f"Assumptions!$B${INGR_ROWS['Green Gram (Moong)']}",
    "Horse Gram":      f"Assumptions!$B${INGR_ROWS['Horse Gram (Kollu)']}",
    "Moth Bean":       f"Assumptions!$B${INGR_ROWS['Moth Bean (Naripayiru)']}",
    "Barnyard Millet": f"Assumptions!$B${INGR_ROWS['Barnyard Millet grains']}",
    "Foxtail Millet":  f"Assumptions!$B${INGR_ROWS['Foxtail Millet grains']}",
    "Little Millet":   f"Assumptions!$B${INGR_ROWS['Little Millet grains']}",
    "Pearl Millet":    f"Assumptions!$B${INGR_ROWS['Pearl Millet grains']}",
    "Black Urad Dal":  f"Assumptions!$B${INGR_ROWS['Black Urad Dal']}",
    "Moong Dal":       f"Assumptions!$B${INGR_ROWS['Moong Dal']}",
    "Rice":            f"Assumptions!$B${INGR_ROWS['Rice (raw)']}",
    "Other Ingredient": None,
}

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: LADDU RECIPES
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Laddu Recipes")
LADDU_NAMES = list(LADDU_RECIPES.keys())
N = len(LADDU_NAMES)

ws2.column_dimensions["A"].width = 26
for i in range(N):
    ws2.column_dimensions[get_column_letter(2 + i)].width = 22

hdr(ws2, 1, 1, "SriLaYa — Laddu Recipes (grams per 1 kg of finished product)", span=N+1, height=28, sz=13)
nt(ws2, 2, 1, "Pre-filled recipes from ops app. RED header = recipe needed — fill in yellow cells.")
ws2.merge_cells(f"A2:{get_column_letter(N+1)}2")

hdr(ws2, 3, 1, "Ingredient", G_DARK, WHITE, sz=10)
for j, name in enumerate(LADDU_NAMES):
    has_r = bool(LADDU_RECIPES[name])
    hdr(ws2, 3, 2+j, name, G_MID if has_r else RED_HDR, WHITE, sz=10)
ws2.row_dimensions[3].height = 52

for i, ing in enumerate(ALL_INGREDIENTS):
    r = 4 + i
    lbl(ws2, r, 1, ing, bold=True)
    for j, name in enumerate(LADDU_NAMES):
        grams = LADDU_RECIPES[name].get(ing)
        col = 2 + j
        c = ws2.cell(row=col, column=col)  # will overwrite below
        c = ws2.cell(row=r, column=col, value=grams)
        c.border = cb()
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0'
        if bool(LADDU_RECIPES[name]):
            if grams:
                c.fill = hf(G_LIGHT); c.font = bf(11)
            else:
                c.fill = hf(WHITE); c.font = bf(10, "AAAAAA"); c.value = "—"
        else:
            c.fill = hf(YELLOW); c.font = bf(11, "0000FF", bold=True); c.border = thick_b()

# Total row
total_r = 4 + len(ALL_INGREDIENTS)
lbl(ws2, total_r, 1, "TOTAL (must = 1000g)", bold=True)
ws2.cell(row=total_r, column=1).fill = hf(G_LIGHT)
for j in range(N):
    col = 2 + j
    c = ws2.cell(row=total_r, column=col,
                 value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_r-1})")
    c.number_format = '#,##0'; c.border = cb()
    c.font = bf(11, bold=True); c.fill = hf(G_LIGHT)
    c.alignment = Alignment(horizontal="right", vertical="center")

# ── RM Cost per kg section ────────────────────────────────────────────────────
cost_hdr_r = total_r + 2
section(ws2, cost_hdr_r, "RM Cost per kg — auto-calculated from Assumptions ingredient rates", ncols=N+1)

cost_col_hdr_r = cost_hdr_r + 1
lbl(ws2, cost_col_hdr_r, 1, "Ingredient", bold=True)
ws2.cell(row=cost_col_hdr_r, column=1).fill = hf(G_LIGHT)
for j in range(N):
    c = ws2.cell(row=cost_col_hdr_r, column=2+j, value="₹ cost / kg product")
    c.font = bf(10, bold=True); c.fill = hf(G_LIGHT); c.border = cb()
    c.alignment = Alignment(horizontal="center")

for i, ing in enumerate(ALL_INGREDIENTS):
    recipe_r = 4 + i
    r = cost_col_hdr_r + 1 + i
    lbl(ws2, r, 1, ing)
    rate_ref = ING_RATE.get(ing)
    for j in range(N):
        col = 2 + j
        qty_ref = f"{get_column_letter(col)}{recipe_r}"
        if rate_ref:
            formula = f'=IFERROR(IF(ISNUMBER({qty_ref}),{qty_ref}/1000*{rate_ref},0),0)'
        else:
            formula = '=0'
        c = ws2.cell(row=r, column=col, value=formula)
        c.number_format = '₹#,##0.00'; c.border = cb(); c.fill = hf(WHITE); c.font = bf(11)
        c.alignment = Alignment(horizontal="right", vertical="center")

RM_COST_ROW = cost_col_hdr_r + 1 + len(ALL_INGREDIENTS)
lbl(ws2, RM_COST_ROW, 1, "RM COST per kg of product (₹)", bold=True)
ws2.cell(row=RM_COST_ROW, column=1).fill = hf(AMBER)
for j in range(N):
    col = 2 + j
    start = cost_col_hdr_r + 1
    end = RM_COST_ROW - 1
    c = ws2.cell(row=RM_COST_ROW, column=col,
                 value=f"=SUM({get_column_letter(col)}{start}:{get_column_letter(col)}{end})")
    c.number_format = '₹#,##0.00'; c.border = cb()
    c.font = bf(11, bold=True); c.fill = hf(AMBER)
    c.alignment = Alignment(horizontal="right", vertical="center")

LADDU_COL = {name: 2 + i for i, name in enumerate(LADDU_NAMES)}

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: COST MODEL
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cost Model")

col_widths3 = [20, 30, 12, 8, 8, 12, 13, 11, 11, 10, 11, 10, 13, 14, 14, 14, 16, 12]
for i, w in enumerate(col_widths3):
    ws3.column_dimensions[get_column_letter(i+1)].width = w

hdr(ws3, 1, 1, "SriLaYa — Per-Variant Cost Model  (199 active variants)", span=18, height=28, sz=13)
nt(ws3, 2, 1,
   "Col F: RM rate/kg (BLUE = from Laddu Recipes sheet, YELLOW = fill in).  "
   "M: LANDED COST.  N/O/P: Suggested price at 35%/45%/55% margin.  Q: YOUR FINAL PRICE (edit column Q).  S: Actual profit % on Q.")
ws3.merge_cells("A2:R2")

hdrs3 = [
    "Category", "Product", "SKU", "Size", "Wt (g)",
    "RM Rate\n(₹/kg)", "RM Cost\n/unit (₹)",
    "Freight\n/unit", "Pkg\n/unit", "Label\n/unit",
    "Labour\n/unit", "Other\n/unit",
    "LANDED\nCOST (₹)",
    "@ 35%\nMargin (₹)", "@ 45%\nMargin (₹)", "@ 55%\nMargin (₹)",
    "YOUR FINAL\nPRICE (₹)",
    "Actual\nProfit %"
]
for col, h in enumerate(hdrs3, 1):
    hdr(ws3, 3, col, h, G_DARK, WHITE, sz=10, height=44)
ws3.freeze_panes = "A4"

# Packaging row refs
PKG_POUCH = {"200g": 11, "250g": 12, "500g": 13, "1kg": 14, "2kg": 15}
PKG_BOX   = {"250g": 19, "500g": 20, "1kg": 21}

data = json.load(open("scripts/data/pricing_data.json", encoding="utf-8"))

row = 4
for product in data:
    cat   = product["category"]["name"] if product["category"] else "Uncategorised"
    title = product["title"]
    is_laddu = "Laddu" in title

    for variant in product["variants"]:
        size  = variant["size"]
        wt_g  = variant["weightGrams"] or 500
        price = float(variant["price"])

        row_fill = hf(G_LIGHT if row % 2 == 0 else WHITE)

        def wc(col, val=None, formula=None, fmt=None, fill=None, fnt=None, align="left"):
            c = ws3.cell(row=row, column=col, value=formula if formula else val)
            c.fill = fill or row_fill
            c.font = fnt or bf(11)
            c.border = cb()
            c.alignment = Alignment(horizontal=align, vertical="center")
            if fmt: c.number_format = fmt
            return c

        wc(1, cat); wc(2, title); wc(3, variant["sku"]); wc(4, size); wc(5, wt_g, align="right")

        F = get_column_letter(6)
        E = get_column_letter(5)

        # F: RM rate/kg
        if is_laddu:
            lcol = LADDU_COL.get(title)
            if lcol:
                rm_formula = f"='Laddu Recipes'!{get_column_letter(lcol)}{RM_COST_ROW}"
                c_f = ws3.cell(row=row, column=6, value=rm_formula)
                c_f.number_format = '₹#,##0.00'; c_f.border = cb()
                c_f.fill = hf(BLUE_T); c_f.font = bf(11)
                c_f.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c_f = ws3.cell(row=row, column=6)
                c_f.fill = hf(YELLOW); c_f.font = bf(11, "0000FF", bold=True)
                c_f.border = thick_b()
                c_f.alignment = Alignment(horizontal="right", vertical="center")
        else:
            sku_key = variant.get("sku", "")
            rm = SKU_RATES.get(sku_key) or em_rate(title)
            c_f = ws3.cell(row=row, column=6, value=float(rm) if rm else None)
            c_f.number_format = '₹#,##0.00'; c_f.border = cb()
            c_f.alignment = Alignment(horizontal="right", vertical="center")
            if rm:
                c_f.fill = row_fill; c_f.font = bf(11)
            else:
                c_f.fill = hf(YELLOW); c_f.font = bf(11, "0000FF", bold=True); c_f.border = thick_b()

        wastage = "Assumptions!$B$28" if is_laddu else "Assumptions!$B$27"

        # G: RM cost/unit  (wastage now stored as decimal, no /100 needed)
        wc(7, formula=f'=IFERROR({F}{row}*{E}{row}/1000*(1+{wastage}),0)',
           fmt='₹#,##0.00', align="right")

        # H: Freight (zero for laddu — local ingredients)
        if is_laddu:
            wc(8, val=0, fmt='₹#,##0.00', align="right")
        else:
            wc(8, formula=f'=(Assumptions!$B$5+Assumptions!$B$6)*{E}{row}/1000',
               fmt='₹#,##0.00', align="right")

        # I: Packaging
        if is_laddu:
            br = PKG_BOX.get(size, 20)
            wc(9, formula=f'=Assumptions!$B${br}', fmt='₹#,##0.00', align="right")
        else:
            pr = PKG_POUCH.get(size, 13)
            wc(9, formula=f'=Assumptions!$B${pr}', fmt='₹#,##0.00', align="right")

        # J: Label
        label_ref = "Assumptions!$B$24" if is_laddu else "Assumptions!$B$23"
        wc(10, formula=f'={label_ref}', fmt='₹#,##0.00', align="right")

        # K: Labour
        if is_laddu:
            wc(11, formula=f'=IFERROR(Assumptions!$B$30/Assumptions!$B$31/1000*{E}{row},0)',
               fmt='₹#,##0.00', align="right")
        else:
            wc(11, formula='=IFERROR(Assumptions!$B$30/Assumptions!$B$32,0)',
               fmt='₹#,##0.00', align="right")

        # L: Other finishing
        other_ref = "Assumptions!$B$26" if is_laddu else "Assumptions!$B$25"
        wc(12, formula=f'={other_ref}', fmt='₹#,##0.00', align="right")

        # M: Landed cost
        G7 = get_column_letter(7)
        wc(13, formula=f'=IFERROR({G7}{row}+H{row}+I{row}+J{row}+K{row}+L{row},"")',
           fmt='₹#,##0.00', align="right",
           fill=hf("DCF0E3"), fnt=bf(11, bold=True))

        # N: Suggested @ 35% conservative margin (B54) — margin stored as decimal
        wc(14, formula=f'=IFERROR(CEILING(M{row}/(1-Assumptions!$B$54),1),"")',
           fmt='₹#,##0.00', align="right",
           fill=hf("E8F5EE"), fnt=bf(11))

        # O: Suggested @ 45% target margin (B55)
        wc(15, formula=f'=IFERROR(CEILING(M{row}/(1-Assumptions!$B$55),1),"")',
           fmt='₹#,##0.00', align="right",
           fill=hf(AMBER), fnt=bf(11, bold=True))

        # P: Suggested @ 55% premium margin (B56)
        wc(16, formula=f'=IFERROR(CEILING(M{row}/(1-Assumptions!$B$56),1),"")',
           fmt='₹#,##0.00', align="right",
           fill=hf("FFF0E0"), fnt=bf(11))

        # Q: Your final price (yellow — edit this column)
        sku_key = variant.get("sku", "")
        q_price = SAVED_PRICES.get(sku_key, price)
        c_o = ws3.cell(row=row, column=17, value=q_price)
        c_o.fill = hf(YELLOW); c_o.font = bf(11, "0000FF", bold=True)
        c_o.border = thick_b()
        c_o.alignment = Alignment(horizontal="right", vertical="center")
        c_o.number_format = '₹#,##0.00'

        # S (col 18): Actual profit % = (Q - M) / Q
        c_s = ws3.cell(row=row, column=18,
                       value=f'=IFERROR((Q{row}-M{row})/Q{row},"")')
        c_s.number_format = '0.0%'
        c_s.alignment = Alignment(horizontal="center", vertical="center")
        c_s.font = bf(11)

        row += 1

out = r"D:\CompanyWebsite\SriLaYa_Cost_Sheet.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Cost Model: {row - 4} variants")
print(f"Laddu Recipes: {sum(1 for v in LADDU_RECIPES.values() if v)} with data, "
      f"{sum(1 for v in LADDU_RECIPES.values() if not v)} need recipes (yellow)")
