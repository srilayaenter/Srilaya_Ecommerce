"""
Generate SriLaYa_Pricing_Review.xlsx from pricing_data.json (run from repo root).
Two sheets:
  1. Pricing Review — one row per variant, editable Price column highlighted in yellow
  2. By Category — pivot summary: product rows grouped under category headers
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open("scripts/data/pricing_data.json", encoding="utf-8"))

# ── colours ──────────────────────────────────────────────────────────────────
GREEN_DARK  = "006A38"
GREEN_LIGHT = "E8F5EE"
AMBER       = "FFF8E1"
YELLOW      = "FFFF00"
GREY_HDR    = "F2F2F2"
WHITE       = "FFFFFF"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def bold(size=11, color="000000"):
    return Font(name="Arial", bold=True, size=size, color=color)

def normal(size=11, color="000000"):
    return Font(name="Arial", size=size, color=color)

# ── Sheet 1: Pricing Review ───────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pricing Review"

headers = [
    "Category", "Product Title", "Product SKU",
    "Variant SKU", "Size", "Weight (g)",
    "Current Price (₹)", "New Price (₹)",
    "GST Rate (%)", "Stock"
]
col_widths = [22, 36, 14, 16, 8, 10, 18, 16, 12, 8]

# Header row
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = bold(11, WHITE)
    cell.fill = hdr_fill(GREEN_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# Data rows
row = 2
for product in data:
    cat = product["category"]["name"] if product["category"] else "Uncategorised"
    for variant in product["variants"]:
        vals = [
            cat,
            product["title"],
            product["sku"],
            variant["sku"],
            variant["size"],
            variant["weightGrams"],
            float(variant["price"]),
            float(variant["price"]),   # New Price pre-filled with current
            float(product["gstRate"]),
            variant["stock"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = normal()
            cell.border = border()
            cell.alignment = Alignment(vertical="center")
            # Highlight editable "New Price" column in yellow
            if col == 8:
                cell.fill = PatternFill("solid", fgColor=YELLOW)
                cell.font = Font(name="Arial", size=11, color="0000FF")
            elif row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
            # Right-align numbers
            if col in (6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col in (7, 8):
                    cell.number_format = '₹#,##0.00'
                if col == 9:
                    cell.number_format = '0.00"%"'
        row += 1

# Legend note below data
ws.cell(row=row + 1, column=1, value="Legend:").font = bold()
ws.cell(row=row + 2, column=1, value="• Yellow (blue text) = New Price column — edit these values, leave others unchanged").font = normal(10)
ws.cell(row=row + 3, column=1, value="• Current Price is pre-filled as the starting point").font = normal(10)

# ── Sheet 2: By Category ──────────────────────────────────────────────────────
ws2 = wb.create_sheet("By Category")

# Group products by category
from collections import defaultdict
by_cat = defaultdict(list)
for product in data:
    cat = product["category"]["name"] if product["category"] else "Uncategorised"
    by_cat[cat].append(product)

headers2 = ["Product", "SKU", "200g", "250g", "500g", "1kg", "2kg", "GST%"]
col_widths2 = [36, 14, 10, 10, 10, 10, 10, 8]

for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = bold(11, WHITE)
    cell.fill = hdr_fill(GREEN_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.row_dimensions[1].height = 24
ws2.freeze_panes = "A2"

SIZE_COLS = {"200g": 3, "250g": 4, "500g": 5, "1kg": 6, "2kg": 7}

row2 = 2
for cat in sorted(by_cat.keys()):
    # Category header
    cat_cell = ws2.cell(row=row2, column=1, value=cat.upper())
    cat_cell.font = bold(10, WHITE)
    cat_cell.fill = hdr_fill("2E6F40")
    ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=8)
    ws2.row_dimensions[row2].height = 18
    row2 += 1

    products = sorted(by_cat[cat], key=lambda p: p["title"])
    for i, product in enumerate(products):
        price_by_size = {v["size"]: float(v["price"]) for v in product["variants"]}
        fill = PatternFill("solid", fgColor=GREEN_LIGHT) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)

        vals2 = [product["title"], product["sku"]]
        for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
            cell = ws2.cell(row=row2, column=col)
            cell.border = border()
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

        ws2.cell(row=row2, column=1, value=product["title"]).font = normal()
        ws2.cell(row=row2, column=2, value=product["sku"]).font = normal()
        ws2.cell(row=row2, column=8, value=float(product["gstRate"])).font = normal()

        for size, col in SIZE_COLS.items():
            if size in price_by_size:
                cell = ws2.cell(row=row2, column=col, value=price_by_size[size])
                cell.number_format = '₹#,##0'
                cell.font = normal()
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                ws2.cell(row=row2, column=col, value="—").font = Font(name="Arial", size=11, color="AAAAAA")
        row2 += 1

    row2 += 1  # blank row between categories

out_path = r"D:\CompanyWebsite\SriLaYa_Pricing_Review.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheet 1 (Pricing Review): {row - 2} variant rows")
print(f"Sheet 2 (By Category): {len(by_cat)} categories, {sum(len(v) for v in by_cat.values())} products")
