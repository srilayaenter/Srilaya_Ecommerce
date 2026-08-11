"""
Evenmore bulk/wholesale rates from evenmorefoods.com/products/all
Tiers scraped: 5-24kg, 25-99kg, 100kg+ (per kg rates)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# All bulk rates scraped from Evenmore (per kg, not per pack)
# Format: { product: { tier1_label: price, tier2_label: price, tier3_label: price } }
BULK = {
    # Millet Flakes (per kg)
    "Foxtail Flakes":         {"5-24kg": 106, "25-99kg": 103, "100kg+": 101},
    "Little Flakes":          {"5-24kg": 191, "25-99kg": 188, "100kg+": 184},
    "Kodo Flakes":            {"5-24kg": 116, "25-99kg": 113, "100kg+": 111},
    "Barnyard Flakes":        {"5-24kg": 197, "25-99kg": 194, "100kg+": 190},
    "Ragi Flakes":            {"5-24kg": 82,  "25-99kg": 79,  "100kg+": 76},
    "Pearl Flakes":           {"5-24kg": 79,  "25-99kg": 76,  "100kg+": 74},
    "White Sorghum Flakes":   {"5-24kg": 80,  "25-99kg": 77,  "100kg+": 75},
    "Red Sorghum Flakes":     {"5-24kg": 88,  "25-99kg": 85,  "100kg+": 83},
    "Wheat Flakes":           {"5-24kg": 94,  "25-99kg": 92,  "100kg+": 90},
    "Barley Flakes":          {"5-24kg": 92,  "25-99kg": 89,  "100kg+": 87},
    "Mapillai Samba Flakes":  {"5-24kg": 95,  "25-99kg": 92,  "100kg+": 90},
    "Karupukavuni Flakes":    {"5-24kg": 147, "25-99kg": 144, "100kg+": 142},
    "Poongar Flakes":         {"5-24kg": 100, "25-99kg": 97,  "100kg+": 95},
    "Kattuyanam Flakes":      {"5-24kg": 100, "25-99kg": 97,  "100kg+": 94},
    "Karunguruvai Flakes":    {"5-24kg": 98,  "25-99kg": 95,  "100kg+": 93},
    "Thanga Samba Flakes":    {"5-24kg": 93,  "25-99kg": 90,  "100kg+": 88},
    "Green Gram Flakes":      {"5-24kg": 151, "25-99kg": 148, "100kg+": 145},
    "Horsegram Flakes":       {"5-24kg": 93,  "25-99kg": 90,  "100kg+": 87},
    "Red Matta Rice Flakes":  {"5-24kg": 89,  "25-99kg": 85,  "100kg+": 82},
    "Red Rice Flakes":        {"5-24kg": 93,  "25-99kg": 90,  "100kg+": 86},

    # Millet Rava (per kg)
    "Foxtail Rava":           {"5-24kg": 105, "25-99kg": 103, "100kg+": 99},
    "Little Rava":            {"5-24kg": 216, "25-99kg": 212, "100kg+": 207},
    "Barnyard Rava":          {"5-24kg": 222, "25-99kg": 217, "100kg+": 212},
    "Finger Rava (Ragi)":     {"5-24kg": 79,  "25-99kg": 77,  "100kg+": 73},
    "Pearl Rava":             {"5-24kg": 70,  "25-99kg": 67,  "100kg+": 63},
    "White Sorghum Rava":     {"5-24kg": 72,  "25-99kg": 69,  "100kg+": 65},
    "Mapillai Samba Rava":    {"5-24kg": 111, "25-99kg": 106, "100kg+": 103},

    # Millet Rice (per kg)
    "Foxtail Millet Grains":  {"5-29kg": 68,  "30-99kg": 66,  "100kg+": 64},
    "Little Millet Grains":   {"5-29kg": 140, "30-99kg": 138, "100kg+": 134},
    "Kodo Millet Grains":     {"5-29kg": 80,  "30-99kg": 77,  "100kg+": 74},
    "Barnyard Millet Grains": {"5-29kg": 143, "30-99kg": 139, "100kg+": 136},
    "Finger Millet Grains":   {"5-29kg": 54,  "30-99kg": 51,  "100kg+": 49},
    "Pearl Millet Grains":    {"5-29kg": 45,  "30-99kg": 43,  "100kg+": 41},
    "White Sorghum Grains":   {"5-29kg": 48,  "30-99kg": 46,  "100kg+": 44},
    "Red Sorghum Grains":     {"5-29kg": 53,  "30-99kg": 50,  "100kg+": 47},

    # Millet Parboiled (per kg)
    "Parboiled Foxtail":      {"5-29kg": 99.65,  "30-99kg": 95.77, "100kg+": 93.24},
    "Parboiled Little":       {"5-29kg": 176.89, "30-99kg": 172.27,"100kg+": 168.24},
    "Parboiled Kodo":         {"5-29kg": 109.94, "30-99kg": 105.97,"100kg+": 103.24},
    "Parboiled Barnyard":     {"5-29kg": 178.95, "30-99kg": 174.31,"100kg+": 170.24},
    "Parboiled Ragi":         {"5-29kg": 80.07,  "30-99kg": 76.39, "100kg+": 74.24},
    "Parboiled Pearl":        {"5-29kg": 70.29,  "30-99kg": 66.70, "100kg+": 64.74},
    "Parboiled White Sorghum":{"5-29kg": 70.80,  "30-99kg": 67.21, "100kg+": 65.24},

    # Millet Flour (per kg)
    "Foxtail Flour":          {"5-24kg": 81,  "25-99kg": 76},
    "Little Flour":           {"5-24kg": 148, "25-99kg": 142},
    "Barnyard Flour":         {"5-24kg": 149, "25-99kg": 145},
    "Finger Flour (Ragi)":    {"5-24kg": 59,  "25-99kg": 55},
    "Pearl Flour":            {"5-24kg": 49,  "25-99kg": 45},
    "White Sorghum Flour":    {"5-24kg": 50,  "25-99kg": 46},
    "Sprouted Ragi Flour":    {"5-24kg": 81,  "25-99kg": 77},
    "Whole Wheat Flour":      {"5-24kg": 51,  "25-99kg": 48},
    "Mapillai Samba Flour":   {"5-24kg": 77,  "25-99kg": 73},

    # Muesli & Granola (per kg)
    "Honey Muesli":           {"5-29kg": 265, "30-99kg": 257, "100kg+": 250},
    "Chocolate Muesli":       {"5-29kg": 298, "30-99kg": 289, "100kg+": 283},
    "Strawberry Muesli":      {"5-29kg": 305, "30-99kg": 297, "100kg+": 292},
    "Fruits Granola":         {"5-29kg": 270, "30-99kg": 262, "100kg+": 254},
    "Mixed Fruit and Nuts Muesli": {"5-29kg": 264, "30-99kg": 259, "100kg+": 256},
    "Sugar Free Muesli":      {"5-29kg": 246, "30-99kg": 240, "100kg+": 236},
    "Sugar Free Stevia Muesli":{"5-29kg": 282,"30-99kg": 277, "100kg+": 274},

    # Sweeteners (per kg)
    "Sugarcane Jaggery":      {"5-30kg": 66,  "31-90kg": 63},
    "Coconut Jaggery":        {"5-30kg": 321, "31-90kg": 315},
    "Palm Jaggery":           {"5-30kg": 328, "31-90kg": 325},

    # Malt & Health Mixes (per kg)
    "Health Mix (Sathu Maavu)":{"5kg+": 170},
    "Ragi Malt Carrot":       {"5-24kg": 214, "25kg+": 200},
    "Ragi Malt Beetroot":     {"5-24kg": 207, "25kg+": 193},

    # Traditional Rice (per kg)
    "Mapillai Samba Rice":    {"5-29kg": 69,  "30-99kg": 65,  "100kg+": 63},
    "Karupu Kavuni Rice":     {"5-29kg": 124, "30-99kg": 121, "100kg+": 118},
    "Poongar Rice":           {"5-29kg": 73,  "30-99kg": 69,  "100kg+": 65},
}

# ── Excel output ──────────────────────────────────────────────────────────────
GREEN_DARK = "006A38"
GREEN_LIGHT = "E8F5EE"
WHITE = "FFFFFF"

def cb():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Evenmore Bulk Rates"

headers = ["Product", "Tier 1 (per kg)", "Tier 1 Price", "Tier 2 (per kg)", "Tier 2 Price", "Tier 3 (per kg)", "Tier 3 Price"]
col_widths = [34, 14, 13, 14, 13, 14, 13]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=GREEN_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = cb()
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 24
ws.freeze_panes = "A2"

row = 2
# Group by category
categories = {
    "Millet Flakes": [k for k in BULK if "Flakes" in k],
    "Millet Rava": [k for k in BULK if "Rava" in k],
    "Millet Rice (Raw)": [k for k in BULK if "Grains" in k],
    "Parboiled Millets": [k for k in BULK if "Parboiled" in k],
    "Millet Flour": [k for k in BULK if "Flour" in k],
    "Muesli & Granola": [k for k in BULK if "Muesli" in k or "Granola" in k],
    "Sweeteners": [k for k in BULK if "Jaggery" in k],
    "Malt & Health Mixes": [k for k in BULK if "Health Mix" in k or "Malt" in k],
    "Traditional Rice": [k for k in BULK if "Rice" in k and "Flakes" not in k and "Grains" not in k and "Parboiled" not in k],
}

for cat, products in categories.items():
    # Category header
    cell = ws.cell(row=row, column=1, value=cat.upper())
    cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cell.fill = PatternFill("solid", fgColor="2E6F40")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

    for i, product in enumerate(sorted(products)):
        tiers = BULK[product]
        tier_list = list(tiers.items())
        fill = PatternFill("solid", fgColor=GREEN_LIGHT if i % 2 == 0 else WHITE)

        ws.cell(row=row, column=1, value=product).font = Font(name="Arial", size=11)
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = cb()

        for t_idx, (tier_label, tier_price) in enumerate(tier_list[:3]):
            col_label = 2 + t_idx * 2
            col_price = 3 + t_idx * 2
            lc = ws.cell(row=row, column=col_label, value=tier_label)
            pc = ws.cell(row=row, column=col_price, value=tier_price)
            for c in [lc, pc]:
                c.fill = fill
                c.border = cb()
                c.font = Font(name="Arial", size=11)
            pc.number_format = '₹#,##0.00'
            pc.alignment = Alignment(horizontal="right")

        # Fill empty tier cols if fewer than 3 tiers
        for t_idx in range(len(tier_list), 3):
            for col in [2 + t_idx*2, 3 + t_idx*2]:
                c = ws.cell(row=row, column=col, value="—")
                c.fill = fill; c.border = cb()
                c.font = Font(name="Arial", size=11, color="AAAAAA")
                c.alignment = Alignment(horizontal="center")
        row += 1
    row += 1

out = r"D:\CompanyWebsite\SriLaYa_Evenmore_BulkRates.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total products with bulk rates: {len(BULK)}")
