"""
Generate SriLaYa_Pricing_Comparison.xlsx
Compares Evenmore (supplier) prices vs current SriLaYa retail prices.
Columns: Category | Product | SKU | Size | Evenmore Price (₹) | SriLaYa Price (₹) | Margin % | New Price (₹)
"""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Evenmore prices scraped from evenmorefoods.com ───────────────────────────
# Format: { "product_key": { "200g": price, "500g": price, "1kg": price } }
# Using the retail pack prices (not bulk/wholesale tiers)
EVENMORE = {
    # Millet Flakes (pack prices)
    "Foxtail Flakes":         {"200g": 31,    "500g": None,  "1kg": 114},
    "Little Flakes":          {"200g": 49,    "500g": None,  "1kg": 199},
    "Kodo Flakes":            {"200g": 33,    "500g": None,  "1kg": 124},
    "Barnyard Flakes":        {"200g": 50,    "500g": None,  "1kg": 205},
    "Ragi Flakes":            {"200g": 26,    "500g": None,  "1kg": 91},
    "Pearl Flakes":           {"200g": 26,    "500g": None,  "1kg": 88},
    "White Sorghum Flakes":   {"200g": 26,    "500g": None,  "1kg": 89},
    "Red Sorghum Flakes":     {"200g": 27,    "500g": None,  "1kg": 96},
    "Wheat Flakes":           {"200g": 29,    "500g": None,  "1kg": 103},
    "Barley Flakes":          {"200g": 28,    "500g": None,  "1kg": 101},
    "Mapillai Samba Flakes":  {"200g": 29,    "500g": None,  "1kg": 103},
    "Karupukavuni Flakes":    {"200g": 38,    "500g": None,  "1kg": 155},
    "Poongar Flakes":         {"200g": 30,    "500g": None,  "1kg": 108},
    "Kattuyanam Flakes":      {"200g": 29,    "500g": None,  "1kg": 108},
    "Karunguruvai Flakes":    {"200g": 29,    "500g": None,  "1kg": 106},
    "Thanga Samba Flakes":    {"200g": 28,    "500g": None,  "1kg": 101},
    "Green Gram Flakes":      {"200g": 40,    "500g": None,  "1kg": 159},
    "Horsegram Flakes":       {"200g": 28,    "500g": None,  "1kg": 101},
    "Red Matta Rice Flakes":  {"200g": 27,    "500g": None,  "1kg": 96},
    "Red Rice Flakes":        {"200g": 28,    "500g": None,  "1kg": 101},
    # Moth Bean Flakes not on Evenmore
    "Moth Beans Flakes":      {"200g": None,  "500g": None,  "1kg": None},

    # Millet Rava (pack prices)
    "Foxtail Rava":           {"500g": 61.4,  "1kg": 113},
    "Little Rava":            {"500g": 117.45,"1kg": 225},
    "Barnyard Rava":          {"500g": 120.1, "1kg": 228},
    "Finger Rava":            {"500g": 48.18, "1kg": 86},
    "Pearl Rava":             {"500g": 43.46, "1kg": 77},
    "White Sorghum Rava":     {"500g": 44.58, "1kg": 79},
    "Mapillai Samba Rava":    {"500g": 62.92, "1kg": 116},

    # Millet Rice (1kg pack prices)
    "Foxtail Millet Grains":  {"1kg": 74},
    "Little Millet Grains":   {"1kg": 145},
    "Kodo Millet Grains":     {"1kg": 86},
    "Barnyard Millet Grains": {"1kg": 148},
    "Finger Millet Grains (Ragi)": {"1kg": 57},
    "Pearl Millet Grains (Bajra)": {"1kg": 48},
    "White Sorghum Grains":   {"1kg": 50},
    "Red Sorghum Grains":     {"1kg": 56},

    # Millet Parboiled (1kg pack prices)
    "Parboiled Foxtail":      {"1kg": 103.15},
    "Parboiled Little":       {"1kg": 181.9},
    "Parboiled Kodo":         {"1kg": 113.65},
    "Parboiled Barnyard":     {"1kg": 184},
    "Parboiled Finger Millets (Ragi)": {"1kg": 83.2},
    "Parboiled Pearl":        {"1kg": 73.23},
    "Parboiled White Sorghum":{"1kg": 73.75},

    # Millet Flour (1kg pack prices)
    "Foxtail Flour":          {"1kg": 86},
    "Little Flour":           {"1kg": 155},
    "Barnyard Flour":         {"1kg": 158},
    "Finger Flour (Ragi)":    {"1kg": 64},
    "Pearl Flour":            {"1kg": 56},
    "White Sorghum Flour":    {"1kg": 57},
    "Sprouted Ragi Flour":    {"1kg": 87},
    "Whole Wheat Flour":      {"1kg": 56},
    "Mapillai Samba Flour":   {"1kg": None},  # not listed as 1kg pack

    # Muesli & Granola (500g and 1kg pack prices)
    "Honey Muesli":           {"500g": 151, "1kg": 271},
    "Chocolate Muesli":       {"500g": 168, "1kg": 308},
    "Strawberry Muesli":      {"500g": 171, "1kg": 313},
    "Fruits Granola":         {"500g": 152, "1kg": 278},
    "Mixed Fruit and Nuts Muesli": {"500g": 151, "1kg": 273},
    "Sugar Free Muesli":      {"500g": 141, "1kg": 254},
    "Sugar Free Stevia Muesli":{"500g": 160,"1kg": 290},

    # Sweeteners (1kg pack prices)
    "Sugarcane Jaggery":      {"1kg": 69},
    "Coconut Jaggery":        {"1kg": 328},
    "Palm Jaggery":           {"1kg": 338},

    # Malt & Health Mixes (500g pack prices)
    "Health Mix (Sathu Maavu)":{"500g": 100},
    "Ragi Malt Carrot":       {"500g": 120},
    "Ragi Malt Beetroot":     {"500g": 118},

    # Traditional Rice (1kg pack prices)
    "Mapillai Samba Rice":    {"1kg": 76},
    "Karupu Kavuni Rice":     {"1kg": 130},
    "Poongar Rice":           {"1kg": 80},
    # Seeraga Samba not listed on Evenmore
    "Seeraga Samba Rice":     {"1kg": None},
}

# ── helpers ───────────────────────────────────────────────────────────────────
GREEN_DARK  = "006A38"
GREEN_LIGHT = "E8F5EE"
YELLOW      = "FFFF00"
RED_LIGHT   = "FFDEDE"
WHITE       = "FFFFFF"

def cell_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_fill(c): return PatternFill("solid", fgColor=c)
def bold(size=11, color="000000"): return Font(name="Arial", bold=True, size=size, color=color)
def normal(size=11, color="000000"): return Font(name="Arial", size=size, color=color)

# ── load SriLaYa current prices ───────────────────────────────────────────────
data = json.load(open("scripts/data/pricing_data.json", encoding="utf-8"))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pricing Comparison"

headers = [
    "Category", "Product", "SKU",
    "Variant SKU", "Size",
    "Evenmore Cost (₹)", "SriLaYa Current (₹)", "Margin %",
    "New Retail Price (₹)", "Notes"
]
col_widths = [22, 34, 14, 16, 8, 18, 20, 10, 18, 30]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = bold(11, WHITE)
    cell.fill = hdr_fill(GREEN_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border()
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

row = 2
for product in data:
    cat = product["category"]["name"] if product["category"] else "Uncategorised"
    title = product["title"]

    # Find Evenmore entry — try exact match first, then partial
    em = EVENMORE.get(title)
    if not em:
        for key in EVENMORE:
            if key.lower() in title.lower() or title.lower() in key.lower():
                em = EVENMORE[key]
                break

    for variant in product["variants"]:
        size = variant["size"]
        srilaya_price = float(variant["price"])
        em_price = em.get(size) if em else None

        margin_pct = None
        if em_price and em_price > 0:
            margin_pct = round((srilaya_price - em_price) / em_price * 100, 1)

        # Flag if margin < 20% or no Evenmore price found
        note = ""
        if em_price is None:
            note = "Not on Evenmore"
        elif margin_pct is not None and margin_pct < 20:
            note = f"⚠ Low margin ({margin_pct}%)"
        elif margin_pct is not None and margin_pct < 0:
            note = f"❌ Below cost!"

        row_fill = PatternFill("solid", fgColor=GREEN_LIGHT if row % 2 == 0 else WHITE)
        if note.startswith("❌"):
            row_fill = PatternFill("solid", fgColor=RED_LIGHT)
        elif note.startswith("⚠"):
            row_fill = PatternFill("solid", fgColor="FFF3CD")

        vals = [
            cat, title, product["sku"],
            variant["sku"], size,
            em_price, srilaya_price, margin_pct,
            srilaya_price,  # New price pre-filled with current
            note
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = normal()
            cell.fill = row_fill
            cell.border = cell_border()
            cell.alignment = Alignment(vertical="center")
            if col in (6, 7, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '₹#,##0.00'
            if col == 8 and val is not None:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.0"%"'
            if col == 9:
                cell.fill = PatternFill("solid", fgColor=YELLOW)
                cell.font = Font(name="Arial", size=11, color="0000FF")
        row += 1

# Legend
ws.cell(row=row+1, column=1, value="Legend:").font = bold()
notes = [
    "• Yellow (blue text) = New Retail Price column — edit these values",
    "• 🟡 Yellow rows = margin < 20% — review pricing",
    "• 🔴 Red rows = SriLaYa price below Evenmore cost — fix urgently",
    "• 'Not on Evenmore' = product sourced elsewhere or new — set price manually",
    "• Evenmore Cost = what SriLaYa pays supplier (retail pack price, not bulk rate)",
]
for i, n in enumerate(notes):
    ws.cell(row=row+2+i, column=1, value=n).font = normal(10)

out = r"D:\CompanyWebsite\SriLaYa_Pricing_Comparison.xlsx"
wb.save(out)

# Summary stats
margins = []
no_cost = 0
low_margin = 0
below_cost = 0
for product in data:
    title = product["title"]
    em = EVENMORE.get(title)
    if not em:
        for key in EVENMORE:
            if key.lower() in title.lower() or title.lower() in key.lower():
                em = EVENMORE[key]; break
    for v in product["variants"]:
        ep = em.get(v["size"]) if em else None
        sp = float(v["price"])
        if ep is None:
            no_cost += 1
        else:
            m = (sp - ep) / ep * 100
            margins.append(m)
            if m < 20: low_margin += 1
            if m < 0:  below_cost += 1

print(f"Saved: {out}")
print(f"Total variants: {row-2}")
print(f"No Evenmore cost data: {no_cost}")
print(f"Low margin (<20%): {low_margin}")
print(f"Below cost: {below_cost}")
if margins:
    print(f"Avg margin: {sum(margins)/len(margins):.1f}%  |  Min: {min(margins):.1f}%  |  Max: {max(margins):.1f}%")
