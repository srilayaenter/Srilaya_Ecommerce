"""
SriLaYa Price Sync
Reads column Q (Your Final Price) from the master cost sheet and generates
SQL UPDATE statements for the ProductVariant table.

Usage:
    python scripts/sync-prices.py           → prints SQL to console
    python scripts/sync-prices.py --out prices.sql  → writes to file
"""
import sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from datetime import datetime

EXCEL = r"D:\CompanyWebsite\SriLaYa_Cost_Sheet.xlsx"

def read_prices():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["Cost Model"]
    prices = {}
    skipped = []
    for r in range(4, 300):
        sku = ws.cell(r, 3).value
        q   = ws.cell(r, 17).value
        if not sku:
            continue
        if q is None:
            skipped.append(sku)
            continue
        prices[sku] = float(q)
    wb.close()
    return prices, skipped

def generate_sql(prices):
    lines = [
        f"-- SriLaYa price sync — generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"-- {len(prices)} variants",
        "",
        "BEGIN;",
        "",
    ]
    for sku, price in sorted(prices.items()):
        lines.append(
            f'UPDATE "ProductVariant" SET price = {price:.2f}, "updatedAt" = NOW() WHERE sku = \'{sku}\';'
        )
    lines += [
        "",
        "-- Verify: check any SKUs not found in DB",
        "SELECT sku FROM \"ProductVariant\" WHERE sku IN (",
        "  " + ", ".join(f"'{s}'" for s in sorted(prices.keys())),
        ");",
        "",
        "COMMIT;",
    ]
    return "\n".join(lines)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", help="Write SQL to this file instead of printing")
    args = parser.parse_args()

    prices, skipped = read_prices()
    print(f"Read {len(prices)} prices from Excel", file=sys.stderr)
    if skipped:
        print(f"Skipped (no price set): {', '.join(skipped)}", file=sys.stderr)

    sql = generate_sql(prices)

    out_path = args.out or "scripts/output/prices.sql"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(sql)
    print(f"SQL written to {out_path}", file=sys.stderr)

if __name__ == "__main__":
    main()
